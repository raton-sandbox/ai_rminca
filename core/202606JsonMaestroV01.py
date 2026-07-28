# Harvesting Raton de Minca
# 2026-JN-03
# Ejecutar python D:/AI_RMinca/202606JsonMaestroV01.py
# Archivos Txt de salida
# Archivo excel de entrada D:/AI_RMinca/catalogo_rutas.xlsx

import os
import json
from openpyxl import load_workbook

def generar_json_maestro_con_contingencia(ruta_excel, carpeta_txt, archivo_json_salida):
    print(f"[Proceso] Leyendo archivo Excel desde: {ruta_excel}")
    print(f"[Proceso] Buscando narrativas en el directorio: {carpeta_txt}")
    
    # 1. Cargar el Excel e indexar rutas válidas
    base_excel = {}
    try:
        wb = load_workbook(ruta_excel, data_only=True)
        hoja = wb['CATALOGO']
        
        fila_encabezados = next(hoja.iter_rows(min_row=1, max_row=1, values_only=True))
        columnas = [str(h).strip().upper() for h in fila_encabezados if h]
        
        idx_id_ruta = columnas.index("ID_RUTA")
        idx_perfil = columnas.index("PERFIL_RUTA")
        
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if len(fila) <= idx_id_ruta or fila[idx_id_ruta] is None:
                continue
            
            # CONTROL 1: Saltar las filas que tengan PERFIL_RUTA en blanco
            if len(fila) <= idx_perfil or fila[idx_perfil] is None or str(fila[idx_perfil]).strip() == "":
                continue
            
            id_ruta_excel = str(fila[idx_id_ruta]).strip()
            
            datos_fila = {}
            for i, nombre_columna in enumerate(columnas):
                if i < len(fila):
                    valor = fila[i]
                    #if nombre_columna in ["CONECTA_CON", "INTERESES_TAGS"] and valor:
                    #    datos_fila[nombre_columna.lower()] = [item.strip() for item in str(valor).split(",")]
                    #else:
                    datos_fila[nombre_columna.lower()] = valor if valor is not None else ""
            
            base_excel[id_ruta_excel] = datos_fila
            
        print(f"[Excel] {len(base_excel)} rutas aprobadas para procesamiento.")
    except Exception as e:
        print(f" [!] Error crítico al leer el archivo Excel: {e}")
        return

    # Mapear previamente los archivos TXT que sí existen en la carpeta
    mapa_archivos_txt = {}
    if os.path.exists(carpeta_txt):
        for archivo in os.listdir(carpeta_txt):
            if archivo.endswith(".txt"):
                id_prefijo = archivo[:6].strip().upper()
                mapa_archivos_txt[id_prefijo] = archivo
    else:
        print(f" [AVISO] El directorio de textos '{carpeta_txt}' no existe. Se procesará todo en blanco.")

    # 2. CONTROL 2: El archivo TXT no es obligatorio. Cruzamos usando la base de Excel como maestra.
    diccionario_maestro = {}
    print("[Fusión] Acoplando información narrativa disponible...")
    
    for id_ruta, datos_excel in base_excel.items():
        descripcion_ux = ""
        id_busqueda_txt = id_ruta.upper()
        
        if id_busqueda_txt in mapa_archivos_txt:
            archivo_nombre = mapa_archivos_txt[id_busqueda_txt]
            ruta_completa = os.path.join(carpeta_txt, archivo_nombre)
            
            try:
                with open(ruta_completa, "r", encoding="utf-8") as f:
                    contenido_completo = f.read()
                
                partes = contenido_completo.split("------------------------------------------------------------")
                if len(partes) > 1:
                    descripcion_ux = partes[1].strip()
                print(f"   -> [OK] ID {id_ruta} fusionado con texto.")
            except Exception as e:
                print(f"   -> [AVISO] Error leyendo archivo {archivo_nombre}: {e}. Espacio asignado en blanco.")
        else:
            print(f"   -> [PARCIAL] ID {id_ruta} sin archivo .txt físico. Datos web en blanco.")
        
        diccionario_maestro[id_ruta] = {
            "datos_excel": datos_excel,
            "contenido_web": {
                "descripcion_narrativa": descripcion_ux
            }
        }

    # Asegurar que el directorio de destino exista antes de guardar
    directorio_destino = os.path.dirname(archivo_json_salida)
    if directorio_destino and not os.path.exists(directorio_destino):
        os.makedirs(directorio_destino, exist_ok=True)
        print(f"[Sistema] Creado el directorio de destino: {directorio_destino}")

    # 3. Escritura del archivo JSON definitivo
    try:
        with open(archivo_json_salida, "w", encoding="utf-8") as json_file:
            json.dump(diccionario_maestro, json_file, ensure_ascii=False, indent=4)
        print(f"\n[Éxito] Archivo maestro guardado con éxito en: {archivo_json_salida}")
        print(f"[Fin] Se integraron un total de {len(diccionario_maestro)} rutas.")
    except Exception as e:
        print(f" [!] Error al escribir el archivo JSON de destino: {e}")


# =====================================================================
# BLOCK DE CONFIGURACIÓN DE RUTAS (ORIGEN Y DESTINO)
# =====================================================================
if __name__ == "__main__":
    
    # 1. ORIGEN: Indica dónde está tu archivo Excel de entrada
    ORIGEN_EXCEL = "D:/AI_RMinca/core/catalogo_rutas.xlsx"  # Cambia por la ruta absoluta si está en otra carpeta
    
    # 2. DIRECTORIO: Especifica la carpeta donde están guardados los archivos .txt descargados
    DIRECTORIO_TXT = "D:/AI_RMinca/textos"  # Reemplaza por tu directorio real (ej: "C:/Proyectos/Bot/textos")
    
    # 3. DESTINO: Define el directorio final y el nombre del archivo JSON que se va a generar
    # Puedes guardarlo en la misma carpeta o especificar una ruta de salida completa
    DESTINO_JSON = "D:/AI_RMinca/core/conocimiento_maestro.json"
    
    # Ejecución automática del proceso con los parámetros configurados
    # Rutas excel
    # Carpeta con textos narrativos
    # carpeta para guardar conocimiento_maestro.json
    generar_json_maestro_con_contingencia(
        ruta_excel=ORIGEN_EXCEL,
        carpeta_txt=DIRECTORIO_TXT,
        archivo_json_salida=DESTINO_JSON
    )